"""
SISTEMA HOSPITAL - Proyecto completo
Requisitos implementados:
  - 4 módulos: Pacientes, Médicos, Citas, Medicamentos
  - MySQL con Stored Procedures (CRUD completo)
  - Exportar a Excel (openpyxl) y PDF (fpdf)
  - Validaciones de campos (numérico, email, texto)
  - tkcalendar para selección de fechas
  - Imágenes con Pillow (JPG, PNG, GIF)
  - Temas claro/oscuro
  - Favicon e iconografía en botones
  - Diálogos de confirmación
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import mysql.connector
import re
from PIL import Image, ImageTk
from tkcalendar import DateEntry
from openpyxl import Workbook
from fpdf import FPDF


# ══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN DE TEMAS (claro / oscuro)
# ══════════════════════════════════════════════════════════════

TEMAS = {
    "claro": {
        "bg": "#F5F5F5",
        "fg": "#212121",
        "btn_bg": "#1976D2",
        "btn_fg": "#FFFFFF",
        "entry_bg": "#FFFFFF",
        "header_fg": "#1565C0",
    },
    "oscuro": {
        "bg": "#212121",
        "fg": "#F5F5F5",
        "btn_bg": "#0D47A1",
        "btn_fg": "#FFFFFF",
        "entry_bg": "#424242",
        "header_fg": "#90CAF9",
    },
}

tema_actual = "claro"


def t():
    """Retorna el diccionario del tema activo."""
    return TEMAS[tema_actual]


# ══════════════════════════════════════════════════════════════
#  CONEXIÓN A MYSQL
# ══════════════════════════════════════════════════════════════

def get_conexion():
    """Abre y retorna una conexión nueva a MySQL."""
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="hospital"
    )


# ══════════════════════════════════════════════════════════════
#  UTILIDADES: VALIDACIONES
# ══════════════════════════════════════════════════════════════

def validar_entero(valor, nombre_campo):
    """Verifica que el valor sea un número entero positivo."""
    if not valor.strip().isdigit():
        messagebox.showerror("Error de validación",
                             f"{nombre_campo} solo acepta números enteros.")
        return False
    return True


def validar_texto(valor, nombre_campo, minimo=2, maximo=50):
    """Verifica longitud mínima/máxima y que no tenga caracteres especiales."""
    if len(valor.strip()) < minimo:
        messagebox.showerror("Error de validación",
                             f"{nombre_campo} debe tener al menos {minimo} caracteres.")
        return False
    if len(valor.strip()) > maximo:
        messagebox.showerror("Error de validación",
                             f"{nombre_campo} no puede superar {maximo} caracteres.")
        return False
    if re.search(r"[<>\"';]", valor):
        messagebox.showerror("Error de validación",
                             f"{nombre_campo} contiene caracteres no permitidos.")
        return False
    return True


def validar_email(valor):
    """Valida formato de email con expresión regular."""
    patron = r"^[\w\.-]+@[\w\.-]+\.\w{2,}$"
    if not re.match(patron, valor.strip()):
        messagebox.showerror("Error de validación",
                             "El formato del email no es válido.")
        return False
    return True


def confirmar(mensaje):
    """Muestra diálogo de confirmación antes de operaciones críticas."""
    return messagebox.askyesno("Confirmar operación", mensaje)


# ══════════════════════════════════════════════════════════════
#  UTILIDADES: IMÁGENES (Pillow)
# ══════════════════════════════════════════════════════════════

def cargar_imagen_bytes(ruta):
    """
    Abre una imagen con Pillow, la redimensiona a 150x150
    y la convierte a bytes para guardar en MySQL (LONGBLOB).
    """
    from io import BytesIO
    img = Image.open(ruta)
    img = img.convert("RGB")       # normaliza a RGB (soporta JPG, PNG, GIF)
    img = img.resize((150, 150))   # redimensiona
    buffer = BytesIO()
    img.save(buffer, format="JPEG")
    return buffer.getvalue()       # retorna bytes


def bytes_a_imagetk(datos_bytes, size=(80, 80)):
    """Convierte bytes almacenados en MySQL a una imagen que Tkinter puede mostrar."""
    from io import BytesIO
    img = Image.open(BytesIO(datos_bytes))
    img = img.resize(size)
    return ImageTk.PhotoImage(img)


# ══════════════════════════════════════════════════════════════
#  UTILIDADES: EXPORTACIÓN
# ══════════════════════════════════════════════════════════════

def exportar_excel(titulo, encabezados, filas):
    """Genera un archivo .xlsx con openpyxl."""
    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        title="Guardar Excel"
    )
    if not ruta:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = titulo
    ws.append(encabezados)
    for fila in filas:
        ws.append(list(fila))
    wb.save(ruta)
    messagebox.showinfo("Exportar Excel", f"Guardado en:\n{ruta}")


def exportar_pdf(titulo, encabezados, filas):
    """Genera un archivo .pdf con fpdf con formato de tabla."""
    ruta = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF", "*.pdf")],
        title="Guardar PDF"
    )
    if not ruta:
        return
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, titulo, ln=True, align="C")
    pdf.ln(5)

    ancho = 190 // len(encabezados)

    # Encabezados en negrita
    pdf.set_font("Arial", "B", 10)
    for enc in encabezados:
        pdf.cell(ancho, 8, str(enc), border=1, align="C")
    pdf.ln()

    # Filas de datos
    pdf.set_font("Arial", "", 9)
    for fila in filas:
        for celda in fila:
            pdf.cell(ancho, 7, str(celda), border=1)
        pdf.ln()

    pdf.output(ruta)
    messagebox.showinfo("Exportar PDF", f"Guardado en:\n{ruta}")


# ══════════════════════════════════════════════════════════════
#  MÓDULO PACIENTES
# ══════════════════════════════════════════════════════════════

class ModuloPacientes(tk.Frame):

    def __init__(self, parent):
        super().__init__(parent, bg=t()["bg"])
        self.foto_bytes = None
        self.construir_ui()

    def construir_ui(self):
        tk.Label(self, text="👤  GESTIÓN DE PACIENTES",
                 font=("Arial", 16, "bold"),
                 fg=t()["header_fg"], bg=t()["bg"]).pack(pady=15)

        contenedor = tk.Frame(self, bg=t()["bg"])
        contenedor.pack(fill="both", expand=True, padx=20)       

        # ── Formulario izquierdo ──────────────────────────────
        form = tk.LabelFrame(contenedor, text="Datos del Paciente",
                             bg=t()["bg"], fg=t()["fg"], font=("Arial", 11))
        form.pack(side="left", fill="y", padx=10, pady=5)

        campos = [
            ("ID Paciente:", "id"),
            ("Nombre:",      "nombre"),
            ("Apellido:",    "apellido"),
            ("Teléfono:",    "telefono"),
            ("Email:",       "email"),
        ]

        self.entradas = {}
        for i, (label, key) in enumerate(campos):
            tk.Label(form, text=label, bg=t()["bg"], fg=t()["fg"],
                     font=("Arial", 11)).grid(row=i, column=0, sticky="w", padx=10, pady=6)
            entry = tk.Entry(form, width=25, bg=t()["entry_bg"], fg=t()["fg"])
            entry.grid(row=i, column=1, padx=10, pady=6)
            self.entradas[key] = entry

        # Imagen con Pillow
        tk.Label(form, text="Foto:", bg=t()["bg"], fg=t()["fg"],
                 font=("Arial", 11)).grid(row=5, column=0, sticky="w", padx=10, pady=6)
        self.lbl_foto = tk.Label(form, text="Sin imagen",
                                 bg=t()["entry_bg"], width=15, height=5)
        self.lbl_foto.grid(row=5, column=1, padx=10, pady=6)
        tk.Button(form, text="📷 Seleccionar foto",
                  bg=t()["btn_bg"], fg=t()["btn_fg"],
                  command=self.seleccionar_foto).grid(row=6, column=1, pady=4)

        # Botones CRUD
        btn_frame = tk.Frame(form, bg=t()["bg"])
        btn_frame.grid(row=7, column=0, columnspan=2, pady=10)
        for texto, cmd in [
            ("💾 Guardar",    self.guardar),
            ("✏️ Actualizar", self.actualizar),
            ("🗑️ Eliminar",   self.eliminar),
            ("🧹 Limpiar",    self.limpiar_form),
        ]:
            tk.Button(btn_frame, text=texto, bg=t()["btn_bg"],
                      fg=t()["btn_fg"], width=13,
                      command=cmd).pack(side="left", padx=4)

        # ── Tabla derecha ─────────────────────────────────────
        tabla_frame = tk.Frame(contenedor, bg=t()["bg"])
        tabla_frame.pack(side="left", fill="both", expand=True, padx=10, pady=5)

        cols = ("ID", "Nombre", "Apellido", "Teléfono", "Email")
        self.tabla = ttk.Treeview(tabla_frame, columns=cols,
                                  show="headings", height=15)
        for col in cols:
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=110)
        self.tabla.pack(fill="both", expand=True)
        # Al hacer clic en una fila, se rellena el formulario
        self.tabla.bind("<<TreeviewSelect>>", self.seleccionar_fila)

        # Botones exportar
        exp_frame = tk.Frame(self, bg=t()["bg"])
        exp_frame.pack(pady=8)
        tk.Button(exp_frame, text="📊 Exportar Excel",
                  bg="#388E3C", fg="white",
                  command=self.exportar_excel).pack(side="left", padx=6)
        tk.Button(exp_frame, text="📄 Exportar PDF",
                  bg="#D32F2F", fg="white",
                  command=self.exportar_pdf).pack(side="left", padx=6)

        self.cargar_tabla()

    # ── Imagen ────────────────────────────────────────────────
    def seleccionar_foto(self):
        ruta = filedialog.askopenfilename(
            filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.gif")]
        )
        if ruta:
            self.foto_bytes = cargar_imagen_bytes(ruta)
            img_tk = bytes_a_imagetk(self.foto_bytes)
            self.lbl_foto.config(image=img_tk, text="")
            self.lbl_foto.image = img_tk  # evitar que el garbage collector borre la imagen

    # ── CRUD ──────────────────────────────────────────────────
    def guardar(self):
        if not self._validar():
            return
        try:
            con = get_conexion()
            cur = con.cursor()
            cur.callproc("sp_registrar_paciente", [
                int(self.entradas["id"].get()),
                self.entradas["nombre"].get(),
                self.entradas["apellido"].get(),
                self.entradas["telefono"].get(),
                self.entradas["email"].get(),
                self.foto_bytes,
            ])
            con.commit()
            messagebox.showinfo("Éxito", "Paciente registrado correctamente")
            self.limpiar_form()
            self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def actualizar(self):
        if not self._validar():
            return
        if not confirmar("¿Deseas actualizar este paciente?"):
            return
        try:
            con = get_conexion()
            cur = con.cursor()
            cur.callproc("sp_actualizar_paciente", [
                int(self.entradas["id"].get()),
                self.entradas["nombre"].get(),
                self.entradas["apellido"].get(),
                self.entradas["telefono"].get(),
                self.entradas["email"].get(),
                self.foto_bytes,
            ])
            con.commit()
            messagebox.showinfo("Éxito", "Paciente actualizado")
            self.limpiar_form()
            self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def eliminar(self):
        id_val = self.entradas["id"].get()
        if not validar_entero(id_val, "ID"):
            return
        if not confirmar("¿Estás seguro de eliminar este paciente?"):
            return
        try:
            con = get_conexion()
            cur = con.cursor()
            cur.callproc("sp_eliminar_paciente", [int(id_val)])
            con.commit()
            messagebox.showinfo("Éxito", "Paciente eliminado")
            self.limpiar_form()
            self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def cargar_tabla(self):
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        try:
            con = get_conexion()
            cur = con.cursor()
            cur.callproc("sp_mostrar_pacientes")
            for result in cur.stored_results():
                for fila in result.fetchall():
                    self.tabla.insert("", "end", values=fila) 
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def seleccionar_fila(self, event):
        sel = self.tabla.selection()
        if sel:
            valores = self.tabla.item(sel[0])["values"]
            for i, k in enumerate(["id", "nombre", "apellido", "telefono", "email"]):
                self.entradas[k].delete(0, tk.END)
                self.entradas[k].insert(0, valores[i])

    def limpiar_form(self):
        for entry in self.entradas.values():
            entry.delete(0, tk.END)
        self.foto_bytes = None
        self.lbl_foto.config(image="", text="Sin imagen")

    def _validar(self):
        e = self.entradas
        return (validar_entero(e["id"].get(), "ID") and
                validar_texto(e["nombre"].get(), "Nombre") and
                validar_texto(e["apellido"].get(), "Apellido") and
                validar_entero(e["telefono"].get(), "Teléfono") and
                validar_email(e["email"].get()))

    def exportar_excel(self):
        filas = [self.tabla.item(r)["values"] for r in self.tabla.get_children()]
        exportar_excel("Pacientes",
                       ["ID", "Nombre", "Apellido", "Teléfono", "Email"], filas)

    def exportar_pdf(self):
        filas = [self.tabla.item(r)["values"] for r in self.tabla.get_children()]
        exportar_pdf("Listado de Pacientes",
                     ["ID", "Nombre", "Apellido", "Teléfono", "Email"], filas)


# ══════════════════════════════════════════════════════════════
#  MÓDULO MÉDICOS
# ══════════════════════════════════════════════════════════════

class ModuloMedicos(tk.Frame):

    def __init__(self, parent):
        super().__init__(parent, bg=t()["bg"])
        self.foto_bytes = None
        self.construir_ui()

    def construir_ui(self):
        tk.Label(self, text="🩺  GESTIÓN DE MÉDICOS",
                 font=("Arial", 16, "bold"),
                 fg=t()["header_fg"], bg=t()["bg"]).pack(pady=15)

        contenedor = tk.Frame(self, bg=t()["bg"])
        contenedor.pack(fill="both", expand=True, padx=20)

        form = tk.LabelFrame(contenedor, text="Datos del Médico",
                             bg=t()["bg"], fg=t()["fg"], font=("Arial", 11))
        form.pack(side="left", fill="y", padx=10, pady=5)

        campos = [
            ("ID Médico:",     "id"),
            ("Nombre:",        "nombre"),
            ("Especialidad:",  "especialidad"),
            ("Email:",         "email"),
        ]
        self.entradas = {}
        for i, (label, key) in enumerate(campos):
            tk.Label(form, text=label, bg=t()["bg"], fg=t()["fg"],
                     font=("Arial", 11)).grid(row=i, column=0, sticky="w", padx=10, pady=6)
            entry = tk.Entry(form, width=25, bg=t()["entry_bg"], fg=t()["fg"])
            entry.grid(row=i, column=1, padx=10, pady=6)
            self.entradas[key] = entry

        tk.Label(form, text="Foto:", bg=t()["bg"], fg=t()["fg"],
                 font=("Arial", 11)).grid(row=4, column=0, sticky="w", padx=10)
        self.lbl_foto = tk.Label(form, text="Sin imagen",
                                 bg=t()["entry_bg"], width=15, height=5)
        self.lbl_foto.grid(row=4, column=1, padx=10, pady=6)
        tk.Button(form, text="📷 Seleccionar foto",
                  bg=t()["btn_bg"], fg=t()["btn_fg"],
                  command=self.seleccionar_foto).grid(row=5, column=1, pady=4)

        btn_frame = tk.Frame(form, bg=t()["bg"])
        btn_frame.grid(row=6, column=0, columnspan=2, pady=10)
        for texto, cmd in [
            ("💾 Guardar",    self.guardar),
            ("✏️ Actualizar", self.actualizar),
            ("🗑️ Eliminar",   self.eliminar),
            ("🧹 Limpiar",    self.limpiar_form),
        ]:
            tk.Button(btn_frame, text=texto, bg=t()["btn_bg"],
                      fg=t()["btn_fg"], width=13,
                      command=cmd).pack(side="left", padx=4)

        tabla_frame = tk.Frame(contenedor, bg=t()["bg"])
        tabla_frame.pack(side="left", fill="both", expand=True, padx=10)

        cols = ("ID", "Nombre", "Especialidad", "Email")
        self.tabla = ttk.Treeview(tabla_frame, columns=cols,
                                  show="headings", height=15)
        for col in cols:
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=130)
        self.tabla.pack(fill="both", expand=True)
        self.tabla.bind("<<TreeviewSelect>>", self.seleccionar_fila)

        exp_frame = tk.Frame(self, bg=t()["bg"])
        exp_frame.pack(pady=8)
        tk.Button(exp_frame, text="📊 Exportar Excel",
                  bg="#388E3C", fg="white",
                  command=self.exportar_excel).pack(side="left", padx=6)
        tk.Button(exp_frame, text="📄 Exportar PDF",
                  bg="#D32F2F", fg="white",
                  command=self.exportar_pdf).pack(side="left", padx=6)

        self.cargar_tabla()

    def seleccionar_foto(self):
        ruta = filedialog.askopenfilename(
            filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.gif")])
        if ruta:
            self.foto_bytes = cargar_imagen_bytes(ruta)
            img_tk = bytes_a_imagetk(self.foto_bytes)
            self.lbl_foto.config(image=img_tk, text="")
            self.lbl_foto.image = img_tk

    def guardar(self):
        if not self._validar(): return
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_registrar_medico", [
                int(self.entradas["id"].get()),
                self.entradas["nombre"].get(),
                self.entradas["especialidad"].get(),
                self.entradas["email"].get(),
                self.foto_bytes,
            ])
            con.commit()
            messagebox.showinfo("Éxito", "Médico registrado")
            self.limpiar_form(); self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def actualizar(self):
        if not self._validar(): return
        if not confirmar("¿Actualizar este médico?"): return
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_actualizar_medico", [
                int(self.entradas["id"].get()),
                self.entradas["nombre"].get(),
                self.entradas["especialidad"].get(),
                self.entradas["email"].get(),
                self.foto_bytes,
            ])
            con.commit()
            messagebox.showinfo("Éxito", "Médico actualizado")
            self.limpiar_form(); self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def eliminar(self):
        id_val = self.entradas["id"].get()
        if not validar_entero(id_val, "ID"): return
        if not confirmar("¿Eliminar este médico?"): return
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_eliminar_medico", [int(id_val)])
            con.commit()
            messagebox.showinfo("Éxito", "Médico eliminado")
            self.limpiar_form(); self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def cargar_tabla(self):
        for row in self.tabla.get_children(): self.tabla.delete(row)
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_mostrar_medicos")
            for result in cur.stored_results():
                for fila in result.fetchall():
                    self.tabla.insert("", "end", values=fila)
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def seleccionar_fila(self, event):
        sel = self.tabla.selection()
        if sel:
            valores = self.tabla.item(sel[0])["values"]
            for i, k in enumerate(["id", "nombre", "especialidad", "email"]):
                self.entradas[k].delete(0, tk.END)
                self.entradas[k].insert(0, valores[i])

    def limpiar_form(self):
        for entry in self.entradas.values(): entry.delete(0, tk.END)
        self.foto_bytes = None
        self.lbl_foto.config(image="", text="Sin imagen")

    def _validar(self):
        e = self.entradas
        return (validar_entero(e["id"].get(), "ID") and
                validar_texto(e["nombre"].get(), "Nombre") and
                validar_texto(e["especialidad"].get(), "Especialidad") and
                validar_email(e["email"].get()))

    def exportar_excel(self):
        filas = [self.tabla.item(r)["values"] for r in self.tabla.get_children()]
        exportar_excel("Médicos", ["ID", "Nombre", "Especialidad", "Email"], filas)

    def exportar_pdf(self):
        filas = [self.tabla.item(r)["values"] for r in self.tabla.get_children()]
        exportar_pdf("Listado de Médicos", ["ID", "Nombre", "Especialidad", "Email"], filas)


# ══════════════════════════════════════════════════════════════
#  MÓDULO CITAS
# ══════════════════════════════════════════════════════════════

class ModuloCitas(tk.Frame):

    def __init__(self, parent):
        super().__init__(parent, bg=t()["bg"])
        self.construir_ui()

    def construir_ui(self):
        tk.Label(self, text="📅  GESTIÓN DE CITAS",
                 font=("Arial", 16, "bold"),
                 fg=t()["header_fg"], bg=t()["bg"]).pack(pady=15)

        contenedor = tk.Frame(self, bg=t()["bg"])
        contenedor.pack(fill="both", expand=True, padx=20)

        form = tk.LabelFrame(contenedor, text="Datos de la Cita",
                             bg=t()["bg"], fg=t()["fg"], font=("Arial", 11))
        form.pack(side="left", fill="y", padx=10, pady=5)

        tk.Label(form, text="ID Paciente:", bg=t()["bg"], fg=t()["fg"],
                 font=("Arial", 11)).grid(row=0, column=0, sticky="w", padx=10, pady=8)
        self.id_paciente = tk.Entry(form, width=25, bg=t()["entry_bg"], fg=t()["fg"])
        self.id_paciente.grid(row=0, column=1, padx=10)

        tk.Label(form, text="ID Médico:", bg=t()["bg"], fg=t()["fg"],
                 font=("Arial", 11)).grid(row=1, column=0, sticky="w", padx=10, pady=8)
        self.id_medico = tk.Entry(form, width=25, bg=t()["entry_bg"], fg=t()["fg"])
        self.id_medico.grid(row=1, column=1, padx=10)

        # tkcalendar: campo de fecha con calendario flotante
        tk.Label(form, text="Fecha:", bg=t()["bg"], fg=t()["fg"],
                 font=("Arial", 11)).grid(row=2, column=0, sticky="w", padx=10, pady=8)
        self.fecha = DateEntry(form, width=23, background="#1976D2",
                               foreground="white", borderwidth=2,
                               date_pattern="yyyy-mm-dd")
        self.fecha.grid(row=2, column=1, padx=10)

        # Filtro de fechas para exportación
        tk.Label(form, text="── Filtro exportación ──",
                 bg=t()["bg"], fg=t()["fg"],
                 font=("Arial", 9)).grid(row=3, column=0, columnspan=2, pady=(12, 2))

        tk.Label(form, text="Desde:", bg=t()["bg"],
                 fg=t()["fg"]).grid(row=4, column=0, sticky="w", padx=10)
        self.fecha_desde = DateEntry(form, width=23, date_pattern="yyyy-mm-dd")
        self.fecha_desde.grid(row=4, column=1, padx=10, pady=4)

        tk.Label(form, text="Hasta:", bg=t()["bg"],
                 fg=t()["fg"]).grid(row=5, column=0, sticky="w", padx=10)
        self.fecha_hasta = DateEntry(form, width=23, date_pattern="yyyy-mm-dd")
        self.fecha_hasta.grid(row=5, column=1, padx=10, pady=4)

        btn_frame = tk.Frame(form, bg=t()["bg"])
        btn_frame.grid(row=6, column=0, columnspan=2, pady=10)
        for texto, cmd in [
            ("💾 Agendar",  self.guardar),
            ("🗑️ Eliminar", self.eliminar),
            ("🧹 Limpiar",  self.limpiar_form),
        ]:
            tk.Button(btn_frame, text=texto, bg=t()["btn_bg"],
                      fg=t()["btn_fg"], width=13,
                      command=cmd).pack(side="left", padx=4)

        tabla_frame = tk.Frame(contenedor, bg=t()["bg"])
        tabla_frame.pack(side="left", fill="both", expand=True, padx=10)

        cols = ("ID Cita", "Paciente", "Apellido", "Médico", "Fecha")
        self.tabla = ttk.Treeview(tabla_frame, columns=cols,
                                  show="headings", height=15)
        for col in cols:
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=110)
        self.tabla.pack(fill="both", expand=True)

        exp_frame = tk.Frame(self, bg=t()["bg"])
        exp_frame.pack(pady=8)
        tk.Button(exp_frame, text="📊 Excel (filtro fecha)",
                  bg="#388E3C", fg="white",
                  command=self.exportar_excel).pack(side="left", padx=6)
        tk.Button(exp_frame, text="📄 PDF (filtro fecha)",
                  bg="#D32F2F", fg="white",
                  command=self.exportar_pdf).pack(side="left", padx=6)

        self.cargar_tabla()

    def guardar(self):
        if not validar_entero(self.id_paciente.get(), "ID Paciente"): return
        if not validar_entero(self.id_medico.get(), "ID Médico"): return
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_registrar_cita", [
                int(self.id_paciente.get()),
                int(self.id_medico.get()),
                self.fecha.get(),
            ])
            con.commit()
            messagebox.showinfo("Éxito", "Cita agendada correctamente")
            self.limpiar_form(); self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def eliminar(self):
        sel = self.tabla.selection()
        if not sel:
            messagebox.showwarning("Atención", "Selecciona una cita de la tabla")
            return
        id_cita = self.tabla.item(sel[0])["values"][0]
        if not confirmar(f"¿Eliminar la cita #{id_cita}?"): return
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_eliminar_cita", [id_cita])
            con.commit()
            messagebox.showinfo("Éxito", "Cita eliminada")
            self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def cargar_tabla(self):
        for row in self.tabla.get_children(): self.tabla.delete(row)
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_mostrar_citas")
            for result in cur.stored_results():
                for fila in result.fetchall():
                    self.tabla.insert("", "end", values=fila)
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def limpiar_form(self):
        self.id_paciente.delete(0, tk.END)
        self.id_medico.delete(0, tk.END)

    def _filas_filtradas(self):
        """Devuelve solo las citas dentro del rango de fechas del filtro."""
        desde = self.fecha_desde.get()
        hasta = self.fecha_hasta.get()
        return [
            self.tabla.item(r)["values"]
            for r in self.tabla.get_children()
            if desde <= str(self.tabla.item(r)["values"][4]) <= hasta
        ]

    def exportar_excel(self):
        exportar_excel("Citas",
                       ["ID Cita", "Paciente", "Apellido", "Médico", "Fecha"],
                       self._filas_filtradas())

    def exportar_pdf(self):
        exportar_pdf("Listado de Citas",
                     ["ID Cita", "Paciente", "Apellido", "Médico", "Fecha"],
                     self._filas_filtradas())


# ══════════════════════════════════════════════════════════════
#  MÓDULO MEDICAMENTOS
# ══════════════════════════════════════════════════════════════

class ModuloMedicamentos(tk.Frame):

    def __init__(self, parent):
        super().__init__(parent, bg=t()["bg"])
        self.foto_bytes = None
        self.construir_ui()

    def construir_ui(self):
        tk.Label(self, text="💊  GESTIÓN DE MEDICAMENTOS",
                 font=("Arial", 16, "bold"),
                 fg=t()["header_fg"], bg=t()["bg"]).pack(pady=15)

        contenedor = tk.Frame(self, bg=t()["bg"])
        contenedor.pack(fill="both", expand=True, padx=20)

        form = tk.LabelFrame(contenedor, text="Datos del Medicamento",
                             bg=t()["bg"], fg=t()["fg"], font=("Arial", 11))
        form.pack(side="left", fill="y", padx=10, pady=5)

        campos = [
            ("ID:",         "id"),
            ("Nombre:",     "nombre"),
            ("Categoría:",  "categoria"),
            ("Stock:",      "stock"),
        ]
        self.entradas = {}
        for i, (label, key) in enumerate(campos):
            tk.Label(form, text=label, bg=t()["bg"], fg=t()["fg"],
                     font=("Arial", 11)).grid(row=i, column=0, sticky="w", padx=10, pady=6)
            entry = tk.Entry(form, width=25, bg=t()["entry_bg"], fg=t()["fg"])
            entry.grid(row=i, column=1, padx=10, pady=6)
            self.entradas[key] = entry

        tk.Label(form, text="Foto:", bg=t()["bg"], fg=t()["fg"],
                 font=("Arial", 11)).grid(row=4, column=0, sticky="w", padx=10)
        self.lbl_foto = tk.Label(form, text="Sin imagen",
                                 bg=t()["entry_bg"], width=15, height=5)
        self.lbl_foto.grid(row=4, column=1, padx=10, pady=6)
        tk.Button(form, text="📷 Seleccionar foto",
                  bg=t()["btn_bg"], fg=t()["btn_fg"],
                  command=self.seleccionar_foto).grid(row=5, column=1, pady=4)

        # Filtro por categoría
        filtro_frame = tk.Frame(form, bg=t()["bg"])
        filtro_frame.grid(row=6, column=0, columnspan=2, pady=6)
        tk.Label(filtro_frame, text="Filtrar categoría:",
                 bg=t()["bg"], fg=t()["fg"]).pack(side="left")
        self.filtro_cat = tk.Entry(filtro_frame, width=14,
                                   bg=t()["entry_bg"], fg=t()["fg"])
        self.filtro_cat.pack(side="left", padx=4)
        tk.Button(filtro_frame, text="🔍",
                  bg=t()["btn_bg"], fg=t()["btn_fg"],
                  command=self.filtrar_categoria).pack(side="left")

        btn_frame = tk.Frame(form, bg=t()["bg"])
        btn_frame.grid(row=7, column=0, columnspan=2, pady=10)
        for texto, cmd in [
            ("💾 Guardar",    self.guardar),
            ("✏️ Actualizar", self.actualizar),
            ("🗑️ Eliminar",   self.eliminar),
            ("🧹 Limpiar",    self.limpiar_form),
        ]:
            tk.Button(btn_frame, text=texto, bg=t()["btn_bg"],
                      fg=t()["btn_fg"], width=13,
                      command=cmd).pack(side="left", padx=4)

        tabla_frame = tk.Frame(contenedor, bg=t()["bg"])
        tabla_frame.pack(side="left", fill="both", expand=True, padx=10)

        cols = ("ID", "Nombre", "Categoría", "Stock")
        self.tabla = ttk.Treeview(tabla_frame, columns=cols,
                                  show="headings", height=15)
        for col in cols:
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=130)
        self.tabla.pack(fill="both", expand=True)
        self.tabla.bind("<<TreeviewSelect>>", self.seleccionar_fila)

        exp_frame = tk.Frame(self, bg=t()["bg"])
        exp_frame.pack(pady=8)
        tk.Button(exp_frame, text="📊 Exportar Excel",
                  bg="#388E3C", fg="white",
                  command=self.exportar_excel).pack(side="left", padx=6)
        tk.Button(exp_frame, text="📄 Exportar PDF",
                  bg="#D32F2F", fg="white",
                  command=self.exportar_pdf).pack(side="left", padx=6)

        self.cargar_tabla()

    def seleccionar_foto(self):
        ruta = filedialog.askopenfilename(
            filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.gif")])
        if ruta:
            self.foto_bytes = cargar_imagen_bytes(ruta)
            img_tk = bytes_a_imagetk(self.foto_bytes)
            self.lbl_foto.config(image=img_tk, text="")
            self.lbl_foto.image = img_tk

    def guardar(self):
        if not self._validar(): return
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_registrar_medicamento", [
                int(self.entradas["id"].get()),
                self.entradas["nombre"].get(),
                self.entradas["categoria"].get(),
                int(self.entradas["stock"].get()),
                self.foto_bytes,
            ])
            con.commit()
            messagebox.showinfo("Éxito", "Medicamento registrado")
            self.limpiar_form(); self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def actualizar(self):
        if not self._validar(): return
        if not confirmar("¿Actualizar este medicamento?"): return
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_actualizar_medicamento", [
                int(self.entradas["id"].get()),
                self.entradas["nombre"].get(),
                self.entradas["categoria"].get(),
                int(self.entradas["stock"].get()),
                self.foto_bytes,
            ])
            con.commit()
            messagebox.showinfo("Éxito", "Medicamento actualizado")
            self.limpiar_form(); self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def eliminar(self):
        id_val = self.entradas["id"].get()
        if not validar_entero(id_val, "ID"): return
        if not confirmar("¿Eliminar este medicamento?"): return
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_eliminar_medicamento", [int(id_val)])
            con.commit()
            messagebox.showinfo("Éxito", "Medicamento eliminado")
            self.limpiar_form(); self.cargar_tabla()
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def cargar_tabla(self):
        for row in self.tabla.get_children(): self.tabla.delete(row)
        try:
            con = get_conexion(); cur = con.cursor()
            cur.callproc("sp_mostrar_medicamentos")
            for result in cur.stored_results():
                for fila in result.fetchall():
                    self.tabla.insert("", "end", values=fila)
        except mysql.connector.Error as e:
            messagebox.showerror("Error MySQL", str(e))
        finally:
            cur.close(); con.close()

    def filtrar_categoria(self):
        """Filtra la tabla visualmente por categoría sin consultar la BD."""
        cat = self.filtro_cat.get().lower()
        for row in self.tabla.get_children():
            vals = self.tabla.item(row)["values"]
            if cat in str(vals[2]).lower():
                self.tabla.reattach(row, "", "end")
            else:
                self.tabla.detach(row)

    def seleccionar_fila(self, event):
        sel = self.tabla.selection()
        if sel:
            valores = self.tabla.item(sel[0])["values"]
            for i, k in enumerate(["id", "nombre", "categoria", "stock"]):
                self.entradas[k].delete(0, tk.END)
                self.entradas[k].insert(0, valores[i])

    def limpiar_form(self):
        for entry in self.entradas.values(): entry.delete(0, tk.END)
        self.foto_bytes = None
        self.lbl_foto.config(image="", text="Sin imagen")

    def _validar(self):
        e = self.entradas
        return (validar_entero(e["id"].get(), "ID") and
                validar_texto(e["nombre"].get(), "Nombre") and
                validar_texto(e["categoria"].get(), "Categoría") and
                validar_entero(e["stock"].get(), "Stock"))

    def exportar_excel(self):
        filas = [self.tabla.item(r)["values"] for r in self.tabla.get_children()]
        exportar_excel("Medicamentos",
                       ["ID", "Nombre", "Categoría", "Stock"], filas)

    def exportar_pdf(self):
        filas = [self.tabla.item(r)["values"] for r in self.tabla.get_children()]
        exportar_pdf("Listado de Medicamentos",
                     ["ID", "Nombre", "Categoría", "Stock"], filas)


# ══════════════════════════════════════════════════════════════
#  APLICACIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════

class App:

    def __init__(self, root):
        self.root = root
        self.root.title("🏥  SISTEMA HOSPITAL")
        self.root.geometry("1100x650")
        self.root.configure(bg=t()["bg"])

        # Favicon: coloca un archivo favicon.ico en la misma carpeta del script
        try:
            self.root.iconbitmap("favicon.ico")
        except Exception:
            pass  # si no hay favicon, continúa sin error

        self._construir_menu()
        self._construir_notebook()

    def _construir_menu(self):
        """Barra de menú con selector de tema claro/oscuro."""
        menubar = tk.Menu(self.root)
        menu_tema = tk.Menu(menubar, tearoff=0)
        menu_tema.add_command(label="☀️  Tema Claro",
                              command=lambda: self.cambiar_tema("claro"))
        menu_tema.add_command(label="🌙  Tema Oscuro",
                              command=lambda: self.cambiar_tema("oscuro"))
        menubar.add_cascade(label="🎨 Tema", menu=menu_tema)
        self.root.config(menu=menubar)

    def _construir_notebook(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        modulos = {
            "👤 Pacientes":    ModuloPacientes,
            "🩺 Médicos":      ModuloMedicos,
            "📅 Citas":        ModuloCitas,
            "💊 Medicamentos": ModuloMedicamentos,
        }

        for nombre, Clase in modulos.items():
            frame = Clase(self.notebook)
            self.notebook.add(frame, text=nombre)

    def cambiar_tema(self, nuevo_tema):
        """
        Cambia el tema global y reconstruye los módulos
        para que los colores se actualicen en toda la interfaz.
        """
        global tema_actual
        tema_actual = nuevo_tema
        self.notebook.destroy()
        self.root.configure(bg=t()["bg"])
        self._construir_notebook()


# ══════════════════════════════════════════════════════════════
#  PUNTO DE ENTRADA
# ══════════════════════════════════════════════════════════════

root = tk.Tk()
app = App(root)
root.mainloop()
